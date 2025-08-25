import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference

archivo = pd.read_csv('ncr_ride_bookings.csv')
# print(archivo[['Driver Ratings', 'Customer Rating', 'Payment Method']])

tabla = archivo.pivot_table(index='Payment Method', columns='Driver Ratings', values='Customer Rating', aggfunc='sum').round(0)
# print(tabla)

tabla.to_excel('Reporte.xlsx', startrow=4, sheet_name='Reporte') # Para crear un excel

wb = load_workbook('Reporte.xlsx')
pestana = wb['Reporte']

minima_columna = wb.active.min_column
maxima_columna = wb.active.max_column
minima_fila = wb.active.min_row
maxima_fila = wb.active.max_row

print(minima_columna, maxima_columna, minima_fila, maxima_fila)

# Grafico

barchart = BarChart()
data = Reference(pestana, min_col=minima_columna+1, max_col=maxima_columna, min_row=minima_fila, max_row=maxima_fila)
categoria = Reference(pestana, min_col=minima_columna, max_col=maxima_columna, min_row=minima_fila+1, max_row=maxima_fila)

barchart.add_data(data, titles_from_data=True)
barchart.set_categories(categoria)
pestana.add_chart(barchart, 'B12')
barchart.title = 'Reporte de Calificaciones'
barchart.style = 5

wb.save('Reporte.xlsx')



